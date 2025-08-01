import traceback
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from selenium .webdriver.common.keys import Keys
import time
import os
from PIL import ImageGrab
from tqdm import tqdm


def login(user, pwd):
    nameDir = r"C:\Users\Administrator\Desktop\治理后截图保存目录"
    driverPath = r"C:\rpa\安装包\Chrome\chromedriver.exe"
    options = webdriver.ChromeOptions()
    # options.add_argument('--ignore-certificate-errors')  # 忽略证书错误
    # options.add_argument('--ignore-ssl-errors=true')  # 忽略SSL错误
    options.binary_location = r"C:\rpa\安装包\Chrome\Chrome-bin\chrome.exe"
    browser = webdriver.Chrome(executable_path=driverPath, options=options)
    browser.maximize_window()
    browser.get("http://pms.pro.js.sgcc.com.cn:32100/yj-pms-portalui/default.html")
    wait = WebDriverWait(browser, 30)
    wait.until(lambda x: x.find_element_by_xpath('//*[@id="adCode"]')).clear()
    wait.until(lambda x: x.find_element_by_xpath('//*[@id="adCode"]')).send_keys(user)  # wangr8, gaox2
    wait.until(lambda x: x.find_element_by_xpath('//*[@id="pwd"]')).clear()
    wait.until(lambda x: x.find_element_by_xpath('//*[@id="pwd"]')).send_keys(pwd)  # wr3185*!!, pms3.0xmz
    wait.until(lambda x: x.find_element_by_xpath('//*[@id="imageField"]')).click()
    wait.until(lambda x: x.find_element_by_xpath('//*[@class="appInner"]//p[text()="电网一张图"]')).click()
    time.sleep(2)
    browser.switch_to.window(browser.window_handles[-1])
    wait.until(lambda x: x.find_element_by_xpath('//*[@class="currentStyle"]')).click()
    time.sleep(1)
    wait.until(lambda x: x.find_element_by_xpath('//*[@class="map-selector"]//li[1]')).click()
    wait.until(lambda x: x.find_element_by_xpath('//*[@class="icon-logo"]')).click()
    time.sleep(2)
    excel = r"C:\Users\Administrator\Desktop\111\111\江宁已治理.xlsx"
    wb = load_workbook(excel, data_only=True)
    ws = wb.active
    for row in tqdm(range(2, ws.max_row + 1)):
        name = str(ws[f"F{row}"].value).strip().replace('PMS_', '')
        if os.path.exists(os.path.join(nameDir, f'{name}_治理后.png')):
            continue
        try:
            wait.until(lambda x: x.find_element_by_xpath('//*[@placeholder="请输入设备名称"]')).clear()
            wait.until(lambda x: x.find_element_by_xpath('//*[@placeholder="请输入设备名称"]')).send_keys(name)
            wait.until(lambda x: x.find_element_by_xpath('//*[@placeholder="请输入设备名称"]')).send_keys(Keys.ENTER)
            time.sleep(1)

            # serachBtn = wait.until(lambda x: x.find_element_by_xpath('//*[@class="el-icon-search icon-btn"]'))
            # if not serachBtn.get_attribute('style'):
            #     serachBtn.click()
            # else:
            #     serachBtn = wait.until(lambda x: x.find_element_by_xpath('//*[@class="el-icon-close icon-btn"]'))
            #     wait.until(lambda x: x.find_element_by_xpath('//*[@placeholder="请输入设备名称"]')).clear()
            #     wait.until(lambda x: x.find_element_by_xpath('//*[@placeholder="请输入设备名称"]')).send_keys(name)
            #     wait.until(lambda x: x.find_element_by_xpath('//*[@class="el-icon-search icon-btn"]'))
            if not wait.until(lambda x: x.find_element_by_xpath(f'//*[@class="query-box__no-data"]')).get_attribute('style'):
                print(f'{name}, 查询无结果')
                continue
            eles = wait.until(lambda x: x.find_elements_by_xpath('//*[@class="group-header"]'))
            for ele in range(1, len(eles) + 1):
                if '低压台区' not in wait.until(lambda x: x.find_element_by_xpath(f'(//*/div[@class="group-header"])[{ele}]')).text:
                    continue
                if wait.until(lambda x: x.find_element_by_xpath(f'(//*/div[@class="group-list"])[{ele}]')).get_attribute('style'):
                    wait.until(lambda x: x.find_element_by_xpath(f'(//*/div[@class="group-header"])[{ele}]')).click()
                wait.until(lambda x: x.find_element_by_xpath(f'(//*[@class="group-list"])[{ele}]//p[contains(text(), "{name}")]')).click()
                break
                

            time.sleep(6)
            img = ImageGrab.grab(bbox=(0, 115, 1920, 1080))
            img.save(os.path.join(nameDir, f'{name}_治理后.png'))
        except Exception:
            print(traceback.format_exc())
    time.sleep(20)

    return browser


if __name__ ==  "__main__":
    login('gaox2', 'pms3.0xmz')